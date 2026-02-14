"""Excel export functionality with charts and formatting"""

import pandas as pd
from datetime import datetime
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from ..analyzer import StockAnalysis


class ExcelExporter:
    """Export stock analysis to formatted Excel workbooks"""
    
    def export_analysis(self, analyses: List[StockAnalysis], filename: str = None) -> str:
        """
        Export stock analyses to Excel with multiple sheets and charts.
        
        Args:
            analyses: List of StockAnalysis objects
            filename: Output filename (default: analysis_YYYYMMDD.xlsx)
            
        Returns:
            Path to created Excel file
        """
        if filename is None:
            filename = f"analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create summary sheet
        self._create_summary_sheet(wb, analyses)
        
        # Create individual sheets for each stock
        for analysis in analyses:
            self._create_stock_sheet(wb, analysis)
        
        # Save workbook
        wb.save(filename)
        return filename
    
    def _create_summary_sheet(self, wb: Workbook, analyses: List[StockAnalysis]):
        """Create summary sheet with all stocks"""
        ws = wb.create_sheet("Summary", 0)
        
        # Headers
        headers = [
            "Ticker", "Company", "Sector", "Price", "Change %", 
            "RSI", "MACD", "P/E", "Market Cap", "Analyst Recom"
        ]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for analysis in analyses:
            finviz = analysis.finviz_data
            row = [
                analysis.ticker,
                analysis.company_name or "",
                analysis.sector or "",
                f"${analysis.current_price:.2f}" if analysis.current_price else "",
                "",  # Change % - would need historical data
                f"{analysis.rsi:.2f}" if analysis.rsi else "",
                f"{analysis.macd:.2f}" if analysis.macd else "",
                finviz.get('P/E', ''),
                finviz.get('Market Cap', ''),
                finviz.get('Analyst Recom', '')
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_stock_sheet(self, wb: Workbook, analysis: StockAnalysis):
        """Create detailed sheet for a single stock"""
        ws = wb.create_sheet(analysis.ticker)
        
        # Title
        ws['A1'] = f"{analysis.ticker} - {analysis.company_name or 'Stock Analysis'}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Current data section
        row = 3
        ws[f'A{row}'] = "Current Data"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        current_data = [
            ("Price", f"${analysis.current_price:.2f}" if analysis.current_price else "N/A"),
            ("Open", f"${analysis.open:.2f}" if analysis.open else "N/A"),
            ("High", f"${analysis.high:.2f}" if analysis.high else "N/A"),
            ("Low", f"${analysis.low:.2f}" if analysis.low else "N/A"),
            ("Sector", analysis.sector or "N/A"),
            ("Industry", analysis.industry or "N/A"),
        ]
        
        for label, value in current_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Technical indicators
        row += 1
        ws[f'A{row}'] = "Technical Indicators"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        technical_data = [
            ("ATR (14)", f"{analysis.atr:.2f}" if analysis.atr else "N/A"),
            ("EMA 20", f"${analysis.ema20:.2f}" if analysis.ema20 else "N/A"),
            ("EMA 50", f"${analysis.ema50:.2f}" if analysis.ema50 else "N/A"),
            ("EMA 200", f"${analysis.ema200:.2f}" if analysis.ema200 else "N/A"),
            ("RSI", f"{analysis.rsi:.2f}" if analysis.rsi else "N/A"),
            ("MACD", f"{analysis.macd:.2f}" if analysis.macd else "N/A"),
            ("Bollinger Upper", f"${analysis.bollinger_upper:.2f}" if analysis.bollinger_upper else "N/A"),
            ("Bollinger Lower", f"${analysis.bollinger_lower:.2f}" if analysis.bollinger_lower else "N/A"),
        ]
        
        for label, value in technical_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Fundamentals
        row += 1
        ws[f'A{row}'] = "Fundamentals"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        finviz = analysis.finviz_data
        fundamental_data = [
            ("Market Cap", finviz.get('Market Cap', 'N/A')),
            ("P/E Ratio", finviz.get('P/E', 'N/A')),
            ("PEG Ratio", finviz.get('PEG', 'N/A')),
            ("ROE", finviz.get('ROE', 'N/A')),
            ("ROA", finviz.get('ROA', 'N/A')),
            ("Analyst Recom", finviz.get('Analyst Recom', 'N/A')),
            ("Inst Own", finviz.get('Inst Own', 'N/A')),
        ]
        
        for label, value in fundamental_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
    
    def export_watchlist(self, watchlist_data: List[dict], filename: str = None) -> str:
        """Export watchlist to Excel"""
        if filename is None:
            filename = f"watchlist_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        df = pd.DataFrame(watchlist_data)
        df.to_excel(filename, index=False, sheet_name="Watchlist")
        return filename
